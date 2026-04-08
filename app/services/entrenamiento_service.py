"""Servicio de entrenamiento/reentrenamiento del modelo ML."""
import json
import logging
import shutil
from datetime import datetime
from io import BytesIO
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.experimental import enable_iterative_imputer  # noqa: F401
from sklearn.impute import IterativeImputer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score,
    confusion_matrix,
    f1_score,
    precision_score,
    recall_score,
    roc_auc_score,
)
from sklearn.model_selection import RandomizedSearchCV, StratifiedKFold, train_test_split
from sklearn.preprocessing import LabelEncoder, StandardScaler
from xgboost import XGBClassifier

from imblearn.over_sampling import SMOTE
from imblearn.pipeline import Pipeline as ImbPipeline

from app.services.prediccion_service import PrediccionService

logger = logging.getLogger(__name__)

RANDOM_STATE = 42

NUMERIC_COLS = ["Mat", "Rep", "2T", "Prom", "edad"]
CATEGORICAL_COLS = [
    "Grado", "Genero", "Semestre", "Carrera",
    "estrato_socioeconomico", "ocupacion_laboral", "con_quien_vive",
    "apoyo_economico", "modalidad_ingreso", "tipo_colegio",
]
REQUIRED_COLUMNS = NUMERIC_COLS + CATEGORICAL_COLS + ["Abandono"]

# Normalización de acentos (misma que predicciones.py)
_NORMALIZAR = {
    "Carrera": {"Tecnológicas": "Tecnologicas", "No Tecnológicas": "No Tecnologicas"},
    "tipo_colegio": {"Público": "Publico"},
    "modalidad_ingreso": {
        "Prueba de Suficiencia Académica": "Prueba de Suficiencia Academica",
        "Admisión Especial": "Admision Especial",
    },
}


def entrenar_modelo(
    df: pd.DataFrame,
    entrenamiento_id: int,
    db_url_sync: str,
    model_dir: str,
) -> None:
    """Pipeline completo de entrenamiento. Se ejecuta en un thread separado.

    Replica el notebook Abandono_modeling.ipynb:
    1. Validación y limpieza
    2. Label Encoding
    3. IterativeImputer (MICE)
    4. OHE + train/test split + scaling
    5. SMOTE + RandomizedSearchCV (RF, XGBoost, LogReg)
    6. Evaluación y guardado de artefactos candidatos
    """
    from sqlalchemy import create_engine
    from sqlalchemy.orm import Session
    from app.models.entrenamiento_modelo import EntrenamientoModelo

    engine = create_engine(db_url_sync)

    try:
        # --- Actualizar estado a "entrenando" ---
        with Session(engine) as session:
            ent = session.get(EntrenamientoModelo, entrenamiento_id)
            ent.estado = "entrenando"
            session.commit()

        # --- 1. Limpieza ---
        df = df.copy()
        df["Abandono"] = df["Abandono"].astype(str).str.strip().str.lower()
        df = df[df["Abandono"].isin(["si", "no"])].copy()
        if len(df) < 50:
            raise ValueError(f"Se requieren al menos 50 registros válidos, se encontraron {len(df)}")

        for col, mapeo in _NORMALIZAR.items():
            if col in df.columns:
                df[col] = df[col].replace(mapeo)

        # tipo_colegio: valores numéricos → nulo
        if "tipo_colegio" in df.columns:
            df["tipo_colegio"] = df["tipo_colegio"].apply(lambda x: x if isinstance(x, str) else np.nan)

        # modalidad_ingreso: valor 'si' → nulo
        if "modalidad_ingreso" in df.columns:
            df["modalidad_ingreso"] = df["modalidad_ingreso"].replace({"si": np.nan})

        df["Abandono_num"] = (df["Abandono"] == "si").astype(int)

        # --- 2. Label Encoding ---
        label_encoders = {}
        for col in CATEGORICAL_COLS:
            le = LabelEncoder()
            valores_validos = df[col].dropna().unique()
            if len(valores_validos) == 0:
                raise ValueError(f"La columna '{col}' no tiene valores válidos (solo nulos)")
            le.fit(valores_validos)
            df[col + "_enc"] = df[col].apply(
                lambda x, _le=le: _le.transform([x])[0] if pd.notna(x) and x in _le.classes_ else np.nan
            )
            label_encoders[col] = le

        # --- 3. IterativeImputer (MICE) ---
        feature_cols = NUMERIC_COLS + [c + "_enc" for c in CATEGORICAL_COLS]
        X_para_imputar = df[feature_cols].copy()

        iter_imputer = IterativeImputer(
            estimator=RandomForestRegressor(n_estimators=100, random_state=RANDOM_STATE, n_jobs=-1),
            max_iter=10,
            random_state=RANDOM_STATE,
        )
        X_imputado = pd.DataFrame(
            iter_imputer.fit_transform(X_para_imputar),
            columns=feature_cols,
            index=X_para_imputar.index,
        )

        # Post-proceso: redondear categóricas
        for col in CATEGORICAL_COLS:
            enc_col = col + "_enc"
            max_val = len(label_encoders[col].classes_) - 1
            X_imputado[enc_col] = X_imputado[enc_col].round().clip(0, max_val).astype(int)
        X_imputado["edad"] = X_imputado["edad"].round().astype(int)

        # --- 4. Preparar features con OHE ---
        df_ml = pd.DataFrame()
        for v in NUMERIC_COLS:
            df_ml[v] = X_imputado[v]
        for v in CATEGORICAL_COLS:
            df_ml[v] = X_imputado[v + "_enc"]

        df_ml = pd.get_dummies(df_ml, columns=CATEGORICAL_COLS, drop_first=True)

        X = df_ml.copy()
        for col in X.columns:
            if X[col].dropna().isin([0, 1]).all():
                X[col] = X[col].astype(bool)

        y = df["Abandono_num"].copy()
        feature_columns = list(X.columns)

        # --- 5. Train/test split + scaling ---
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=0.2, random_state=RANDOM_STATE, stratify=y
        )

        scaler = StandardScaler()
        X_train_scaled = X_train.copy()
        X_test_scaled = X_test.copy()
        X_train_scaled[NUMERIC_COLS] = scaler.fit_transform(X_train[NUMERIC_COLS])
        X_test_scaled[NUMERIC_COLS] = scaler.transform(X_test[NUMERIC_COLS])

        # --- 6. SMOTE + RandomizedSearchCV ---
        cv_strategy = StratifiedKFold(n_splits=5, shuffle=True, random_state=RANDOM_STATE)
        ratio_desbalance = (y_train == 0).sum() / max((y_train == 1).sum(), 1)

        modelos_config = {
            "Random Forest": {
                "pipeline": ImbPipeline([
                    ("smote", SMOTE(random_state=RANDOM_STATE)),
                    ("modelo", RandomForestClassifier(random_state=RANDOM_STATE, n_jobs=-1)),
                ]),
                "params": {
                    "modelo__n_estimators": [100, 200, 300],
                    "modelo__max_depth": [5, 10, 15, None],
                    "modelo__min_samples_split": [2, 5, 10],
                    "modelo__min_samples_leaf": [1, 2, 4],
                    "modelo__class_weight": ["balanced"],
                },
                "n_iter": 30,
            },
            "XGBoost": {
                "pipeline": ImbPipeline([
                    ("smote", SMOTE(random_state=RANDOM_STATE)),
                    ("modelo", XGBClassifier(random_state=RANDOM_STATE, eval_metric="logloss")),
                ]),
                "params": {
                    "modelo__n_estimators": [100, 200, 300],
                    "modelo__max_depth": [3, 5, 7],
                    "modelo__learning_rate": [0.01, 0.05, 0.1, 0.2],
                    "modelo__min_child_weight": [1, 3, 5],
                    "modelo__subsample": [0.7, 0.8, 0.9],
                    "modelo__colsample_bytree": [0.7, 0.8, 0.9],
                    "modelo__scale_pos_weight": [ratio_desbalance],
                },
                "n_iter": 50,
            },
            "Logistic Regression": {
                "pipeline": ImbPipeline([
                    ("smote", SMOTE(random_state=RANDOM_STATE)),
                    ("modelo", LogisticRegression(max_iter=1000, random_state=RANDOM_STATE)),
                ]),
                "params": {
                    "modelo__C": [0.01, 0.1, 1, 10],
                    "modelo__solver": ["lbfgs", "liblinear"],
                    "modelo__class_weight": ["balanced"],
                },
                "n_iter": 8,  # GridSearch equivalente (pocas combinaciones)
            },
        }

        resultados = {}
        for nombre, config in modelos_config.items():
            logger.info("Entrenando %s para entrenamiento_id=%d", nombre, entrenamiento_id)
            search = RandomizedSearchCV(
                config["pipeline"],
                config["params"],
                n_iter=config["n_iter"],
                scoring="f1",
                cv=cv_strategy,
                random_state=RANDOM_STATE,
                n_jobs=-1,
                refit=True,
            )
            search.fit(X_train_scaled, y_train)

            best_pipeline = search.best_estimator_
            y_pred = best_pipeline.predict(X_test_scaled)
            y_proba = best_pipeline.predict_proba(X_test_scaled)[:, 1]

            metricas = {
                "accuracy": round(accuracy_score(y_test, y_pred), 4),
                "precision": round(precision_score(y_test, y_pred, zero_division=0), 4),
                "recall": round(recall_score(y_test, y_pred, zero_division=0), 4),
                "f1_score": round(f1_score(y_test, y_pred, zero_division=0), 4),
                "roc_auc": round(roc_auc_score(y_test, y_proba), 4),
                "confusion_matrix": confusion_matrix(y_test, y_pred).tolist(),
            }

            resultados[nombre] = {
                "pipeline": best_pipeline,
                "metricas": metricas,
                "best_params": search.best_params_,
            }

        # --- 7. Seleccionar mejor modelo por F1 ---
        mejor_nombre = max(resultados, key=lambda x: resultados[x]["metricas"]["f1_score"])
        mejor = resultados[mejor_nombre]
        mejor_pipeline = mejor["pipeline"]
        mejor_modelo = mejor_pipeline.named_steps["modelo"]

        # --- 8. Guardar artefactos candidatos ---
        candidatos_dir = Path(model_dir) / "candidatos" / str(entrenamiento_id)
        candidatos_dir.mkdir(parents=True, exist_ok=True)

        joblib.dump(mejor_modelo, candidatos_dir / "mejor_modelo.pkl")
        joblib.dump(scaler, candidatos_dir / "scaler.pkl")
        joblib.dump(label_encoders, candidatos_dir / "label_encoders.pkl")
        joblib.dump(iter_imputer, candidatos_dir / "iter_imputer.pkl")
        joblib.dump(feature_columns, candidatos_dir / "feature_columns.pkl")

        # Leer métricas del modelo actual
        metricas_actual = _leer_metricas_modelo_actual(model_dir)

        version = f"v_retrain_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        # --- 9. Actualizar BD con resultados ---
        with Session(engine) as session:
            ent = session.get(EntrenamientoModelo, entrenamiento_id)
            ent.estado = "completado"
            ent.fecha_fin = datetime.now()
            ent.metricas_nuevo = mejor["metricas"]
            ent.metricas_actual = metricas_actual
            ent.parametros_modelo = {
                k.replace("modelo__", ""): (
                    v if not isinstance(v, np.integer) else int(v)
                )
                for k, v in mejor["best_params"].items()
            }
            ent.tipo_mejor_modelo = mejor_nombre
            ent.version_generada = version
            ent.ruta_artefactos_candidatos = str(candidatos_dir)
            session.commit()

        logger.info(
            "Entrenamiento %d completado: mejor=%s, F1=%.4f",
            entrenamiento_id, mejor_nombre, mejor["metricas"]["f1_score"],
        )

    except Exception as exc:
        logger.exception("Error en entrenamiento %d: %s", entrenamiento_id, exc)
        with Session(engine) as session:
            ent = session.get(EntrenamientoModelo, entrenamiento_id)
            ent.estado = "error"
            ent.fecha_fin = datetime.now()
            ent.mensaje_error = str(exc)[:500]
            session.commit()
    finally:
        engine.dispose()


def aceptar_modelo(entrenamiento_id: int, model_dir: str) -> str:
    """Reemplaza el modelo actual con los artefactos candidatos.

    Returns:
        La versión del nuevo modelo.
    """
    model_path = Path(model_dir)
    candidatos_dir = model_path / "candidatos" / str(entrenamiento_id)

    if not candidatos_dir.exists():
        raise FileNotFoundError(f"No se encontraron artefactos candidatos en {candidatos_dir}")

    # Backup del modelo actual
    backup_dir = model_path / "backup"
    backup_dir.mkdir(exist_ok=True)
    for pkl in model_path.glob("*.pkl"):
        shutil.copy2(pkl, backup_dir / pkl.name)

    # Copiar candidatos a producción
    for pkl in candidatos_dir.glob("*.pkl"):
        shutil.copy2(pkl, model_path / pkl.name)

    # Leer versión del entrenamiento
    from sqlalchemy import create_engine
    from sqlalchemy.orm import Session
    from app.core.config import settings
    from app.models.entrenamiento_modelo import EntrenamientoModelo

    engine = create_engine(settings.database_url_sync)
    try:
        with Session(engine) as session:
            ent = session.get(EntrenamientoModelo, entrenamiento_id)
            version = ent.version_generada or "unknown"
            metricas = ent.metricas_nuevo
            tipo = ent.tipo_mejor_modelo
    finally:
        engine.dispose()

    # Actualizar model_info.json
    model_info = {
        "timestamp": datetime.now().isoformat(),
        "version": version,
        "mejor_modelo": tipo,
        "tipo_modelo": tipo,
        "mejor_metricas": metricas,
        "entrenamiento_id": entrenamiento_id,
    }
    with open(model_path / "model_info.json", "w", encoding="utf-8") as f:
        json.dump(model_info, f, ensure_ascii=False, indent=2)

    # Limpiar candidatos
    shutil.rmtree(candidatos_dir, ignore_errors=True)

    # Subir artefactos aceptados a Supabase Storage para persistir en filesystem efímero (Render)
    try:
        from app.model_loader import subir_artefactos_a_supabase
        subir_artefactos_a_supabase(model_dir, settings)
        logger.info("Artefactos del nuevo modelo subidos a Supabase Storage.")
    except Exception as exc:
        logger.warning(
            "No se pudieron subir artefactos a Supabase (el modelo está activo localmente): %s",
            exc,
        )

    return version


def rechazar_modelo(entrenamiento_id: int, model_dir: str) -> None:
    """Elimina los artefactos candidatos sin reemplazar el modelo actual."""
    candidatos_dir = Path(model_dir) / "candidatos" / str(entrenamiento_id)
    if candidatos_dir.exists():
        shutil.rmtree(candidatos_dir, ignore_errors=True)


def recargar_servicio(app_state, model_dir: str) -> None:
    """Recarga el PrediccionService con los nuevos artefactos."""
    app_state.prediccion_service = PrediccionService(model_dir)


def generar_plantilla_excel() -> BytesIO:
    """Genera un Excel plantilla con las columnas requeridas para entrenamiento."""
    wb = Workbook()

    # --- Hoja 1: Datos ---
    ws = wb.active
    ws.title = "Datos"

    headers = NUMERIC_COLS + CATEGORICAL_COLS + ["Abandono"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Filas de ejemplo
    ejemplos = [
        [5, 1, 0, 6.5, 20, "Civil", "Masculino", "Primer", "Tecnologicas",
         "Medio", "No", "Con mis padres", "Total",
         "Curso Preuniversitario/Intensivo", "Publico", "no"],
        [4, 2, 1, 4.8, 22, "Militar", "Femenino", "Segundo", "No Tecnologicas",
         "Bajo", "Si", "Solo/a", "Parcial",
         "Prueba de Suficiencia Academica", "Privado", "si"],
        [6, 0, 0, 7.2, 19, "Civil", "Masculino", "Primer", "Tecnologicas",
         "Alto", "No", "Con Familiares", "Ninguno",
         "Curso Vestibular", "Convenio", "no"],
    ]
    for row_idx, ejemplo in enumerate(ejemplos, 2):
        for col_idx, val in enumerate(ejemplo, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Ajustar anchos
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

    # --- Hoja 2: Instrucciones ---
    ws2 = wb.create_sheet("Instrucciones")
    title_font = Font(bold=True, size=12)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 40

    ws2.cell(row=1, column=1, value="Columna").font = title_font
    ws2.cell(row=1, column=2, value="Descripción").font = title_font
    ws2.cell(row=1, column=3, value="Valores Válidos").font = title_font

    instrucciones = [
        ("Mat", "Materias inscritas", "Número entero (ej: 5)"),
        ("Rep", "Materias reprobadas", "Número entero (ej: 0)"),
        ("2T", "Materias en segunda oportunidad", "Número entero (ej: 1)"),
        ("Prom", "Promedio general", "Número decimal (ej: 6.50)"),
        ("edad", "Edad del estudiante", "Número entero (ej: 20)"),
        ("Grado", "Grado del estudiante", "Civil, Militar"),
        ("Genero", "Género", "Masculino, Femenino"),
        ("Semestre", "Semestre cursado", "Primer, Segundo"),
        ("Carrera", "Tipo de carrera", "Tecnologicas, No Tecnologicas"),
        ("estrato_socioeconomico", "Estrato socioeconómico", "Alto, Medio, Bajo"),
        ("ocupacion_laboral", "¿Trabaja?", "Si, No"),
        ("con_quien_vive", "Con quién vive", "Con mis padres, Con Familiares, Solo/a, Con mi novia, En residencia o alojamiento estudiantil"),
        ("apoyo_economico", "Tipo de apoyo económico", "Total, Parcial, Ninguno"),
        ("modalidad_ingreso", "Modalidad de ingreso", "Curso Preuniversitario/Intensivo, Curso Vestibular, Prueba de Suficiencia Academica, Admision Especial"),
        ("tipo_colegio", "Tipo de colegio de procedencia", "Publico, Privado, Convenio"),
        ("Abandono", "Variable objetivo (a predecir)", "si, no"),
    ]
    for row_idx, (col, desc, valores) in enumerate(instrucciones, 2):
        ws2.cell(row=row_idx, column=1, value=col)
        ws2.cell(row=row_idx, column=2, value=desc)
        ws2.cell(row=row_idx, column=3, value=valores)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def leer_modelo_actual_info(model_dir: str) -> dict | None:
    """Lee model_info.json del modelo en producción."""
    info_path = Path(model_dir) / "model_info.json"
    if info_path.exists():
        with open(info_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def _leer_metricas_modelo_actual(model_dir: str) -> dict | None:
    """Lee las métricas del modelo actual desde model_info.json."""
    info = leer_modelo_actual_info(model_dir)
    if info and "mejor_metricas" in info:
        return info["mejor_metricas"]
    return None
